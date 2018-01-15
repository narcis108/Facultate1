import openpyxl
from django.contrib.auth.models import User
from openpyxl.chart.series import DataPoint
from main_app.database_utils import get_tickets_sold, get_user_tickets
from main_app.models import Film

INITIAL_ROW = 2


def generate_excel_tickets_sold():
    CURRENT_ROW_NUMBER = INITIAL_ROW
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet['A1'].value = 'Film'
    sheet['B1'].value = 'Locuri vandute'
    rezervari = get_tickets_sold()
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    for film in rezervari:
        sheet.cell(row=CURRENT_ROW_NUMBER, column=1).value = Film.objects.get(pk=film['film']).titlu
        sheet.cell(row=CURRENT_ROW_NUMBER, column=2).value = int(film['loc_ocupat__count'])
        CURRENT_ROW_NUMBER += 1

    sheet_chart = wb.create_sheet('Chart')
    labels = openpyxl.chart.Reference(sheet,
                                      min_col=1, min_row=2, max_row=sheet.max_row)
    data = openpyxl.chart.Reference(sheet,
                                    min_col=2, min_row=1, max_row=sheet.max_row)

    chartObj = openpyxl.chart.PieChart()
    chartObj.title = 'Vanzari'
    chartObj.add_data(data, titles_from_data=True)
    chartObj.set_categories(labels)

    sheet_chart.add_chart(chartObj, 'H10')

    wb.save('vanzari1.xlsx')


def generate_excel_users():
    CURRENT_ROW_NUMBER = INITIAL_ROW
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet['A1'].value = 'Utilizator'
    sheet['B1'].value = 'Rezervari efectuate'
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 20
    rezervari = get_user_tickets()
    for rezervare in rezervari:
        sheet.cell(row=CURRENT_ROW_NUMBER, column=1).value = User.objects.get(username=rezervare['persoana']).username
        sheet.cell(row=CURRENT_ROW_NUMBER, column=2).value = int(rezervare['locuri_rezervate__count'])
        CURRENT_ROW_NUMBER += 1
    sheet_chart = wb.create_sheet('Chart')
    labels = openpyxl.chart.Reference(sheet,
                                      min_col=1, min_row=2, max_row=sheet.max_row)
    data = openpyxl.chart.Reference(sheet,
                                    min_col=2, min_row=1, max_row=sheet.max_row)

    chartObj = openpyxl.chart.PieChart()
    chartObj.title = 'Vanzari'
    chartObj.add_data(data, titles_from_data=True)
    chartObj.set_categories(labels)

    sheet_chart.add_chart(chartObj, 'H10')

    wb.save('utilizator.xlsx')
